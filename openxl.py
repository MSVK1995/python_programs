# -*- coding: utf-8 -*-
"""
Created on Wed Jun 13 18:25:41 2018

@author: Vikas Kumar
"""

import openpyxl as xl
from difflib import SequenceMatcher
from openpyxl.styles import PatternFill
import pandas as pd

workbook = xl.load_workbook("company_data_quality_check_Owler.xlsx")
#field_names = []
#field_weight = []
#
#sheet_weightage = workbook["Weightage"]
#
#        
#rows = sheet_weightage.rows
#
#for row in rows:
#    field_names.append(row[0].value)
#    field_weight.append(row[1].value)
#    
    

    
public = workbook["Public"]["D2":"L101"]
index = 0
validate = workbook["Public(Validated)"]["D2":"L101"]


redFill = PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid')


clearFill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')


for (row1, row2) in zip(public, validate):
    for (cell1, cell2) in zip(row1, row2):
        if not cell1.value or cell1.value == "NA":
#            print("Empty", type(cell1.value) , end = "|")
            cell1.fill = redFill
            continue
        else:
            if(isinstance(cell1.value, float) == True or isinstance(cell1.value, int) == True):
#                print("int", type(cell1.value), end = "|")
                if(cell1.value == cell2.value):
                    cell1.fill = clearFill
#                    print(cell1.value, end = "|")
                else:
#                    print("Incorrect int", end = "|")
                    cell1.fill = redFill
            elif (isinstance(cell1.value, str) == True):
#                print("Str", type(cell1.value), end = "|")
#                print(cell1.coordinate, SequenceMatcher(None, cell1.value, cell2.value).ratio())
                if(SequenceMatcher(None, cell1.value, cell2.value).ratio() > 0.35):
                    cell1.fill = clearFill
#                    print("Correct Str", end = "|")
                else:
#                    print("Invalid Str", end = "|")
                    cell1.fill = redFill
                    
#    index+=1
#    if(index == 100):
#        break


#print(public.cell["Name"])
#
#for index, row in enumerate(public.columns):
#    for cell in row:
#        print(public.cell(row=1, column=index + 1).value, cell.value)
#    if index == 5:
#        break
#
#from openpyxl.styles import PatternFill
#
##redFill = PatternFill(start_color='FFEE1111', end_color='FFEE1111', fill_type='solid')
#_cell = public['A1']
#
#print(_cell.value)
#_cell.value = "Name"
#
#print(_cell.value)

#IMPORTANT: SAVE CHANGES MADE in pyfile
                    
                    
public_rating = workbook["Public"]["D2":"N101"]
  
category = {'Address Line 1': 0,
 'City': 1,
 'Country': 3,
 'Employee Count': 9,
 'Founded Date': 8,
 'Phone Number': 5,
 'SIC code/Industry': 6,
 'State': 2,
 'Type (Private/Public)': 7,
 'Zip Code': 4}

ws_dict = pd.read_excel('company_data_quality_check_Owler.xlsx', sheet_name="Weightage")
weight = dict(ws_dict.iloc[:, :].values)

for rating_row in range(0, 100):
    public_rating[rating_row][10].value = 0
                  
for i in range(0, 100):
    for cat_ele in category:
        if public_rating[i][category[cat_ele]].fill.start_color.index == 'FFFFFFFF':
            if(public_rating[i][10].value == 0):
                public_rating[i][10].value = int(weight[cat_ele])
            else:
                public_rating[i][10].value+= int(weight[cat_ele])
    public_rating[i][10].value+= 1
            
workbook.save("company_data_quality_check_Owler.xlsx")